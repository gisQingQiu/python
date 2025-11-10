# -*- coding: utf-8 -*-
"""
Created on Wed Jul 16 19:26:01 2025

@author: Qiu
"""

import os
import numpy as np
import pandas as pd
import rasterio
import geopandas as gpd
import openpyxl
import json
from tqdm import tqdm
from openpyxl.styles import Alignment

# 数据路径
with open(r"制表\母质土种对应表.json", encoding='utf-8') as file: dic = json.load(file)
with open(r"制表\分级标准.json", encoding='utf-8') as file: claparams = json.load(file)
with open(r"制表\字段映射表.json") as file: feild = json.load(file)

# 创建 Excel 工作簿
excel_path = r"制表\表七.xlsx"    #表的存放路径
wb = openpyxl.Workbook()    #创建excel工作簿

for soil in tqdm(claparams.keys(), desc='土壤属性'):
    ws = wb.create_sheet(title=soil)
    unit = claparams[soil][1]
    name = feild[soil]
    point = gpd.read_file(r"样本点.shp").to_crs('EPSG:4527')
    point[(point=='/')|(point=='<空>')] = np.nan
    point.loc[point[name] == 0, name] = np.nan
    # 表头
    ws.append(['母岩母质', '土种类型', '样点统计', ''])
    ws.append(['', '', f'均值/({unit})', '数量/个'])
    ws.merge_cells(start_row=1,start_column=1,end_row=2,end_column=1)
    ws.merge_cells(start_row=1,start_column=2,end_row=2,end_column=2)
    ws.merge_cells(start_row=1,start_column=3,end_row=1,end_column=4)
    
    pos = 3
    total = 0
    for key, values in tqdm(dic.items(), leave=False, desc='母质土种'):
        ws[f'A{pos}'] = key
        ws.merge_cells(start_row=pos,start_column=1,end_row=pos+len(values)-1,end_column=1)
        for value in values:
            ws[f'B{pos}'] = value
            # 统计样点
            taget = point[point['土种_3'] == value].copy()
            taget = taget[name].astype(float)
            taget = taget.dropna()
            if taget.empty:
                ws[f'C{pos}'] = '-'
                ws[f'D{pos}'] = '-'
            else:
                ws[f'C{pos}'] = f'{np.nanmean(taget.values):.1f}'
                ws[f'D{pos}'] = taget.shape[0]
                total += taget.shape[0]
            pos += 1
    point = point[name].dropna().astype(float)
    ws.append(['全县', '-', f'{np.nanmean(point.values):.1f}', total])
    
    uniform_width = 23    # 设置列宽
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = uniform_width
        
    align = Alignment(horizontal='center', vertical='center', wrapText=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align
        
wb.remove(wb["Sheet"])
wb.save(excel_path)


















