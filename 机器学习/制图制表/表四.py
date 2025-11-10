# -*- coding: utf-8 -*-
"""
Created on Tue Apr  1 20:25:44 2025

@author: Qiu
"""

import os
import numpy as np
import pandas as pd
import rasterio
import geopandas as gpd
import openpyxl 
import json
from tqdm import tqdm
from rasterio.mask import mask
from openpyxl.styles import Alignment

# 读取数据
with open(r"制表\分级标准.json", encoding='utf-8') as file: classify = json.load(file)
with open(r"制表\字段映射表.json", encoding='utf-8') as file: feild = json.load(file)
ya = gpd.read_file(r"亚类.shp").dropna(subset=['亚类_3']).set_index('亚类_3').to_crs('EPSG:4527')
point = gpd.read_file(r"样本点.shp").to_crs('EPSG:4527')
point[(point=='/')|(point=='<空>')] = np.nan
# soils = glob(r"栅格数据\*.tif")
soils = r"栅格数据"

# 创建 Excel 工作簿
excel_path = r"制表\表四.xlsx"      #表的存放路径
wb = openpyxl.Workbook()     #创建excel工作簿
ws = wb.active    #打开一个表格

lack = []
for soil, values in tqdm(classify.items(), desc='土壤属性'):
    rasterpath = os.path.join(soils, soil+'.tif')
    if os.path.exists(rasterpath):
        # 创建新表格
        # name = os.path.basename(soil).split('.')[0]
        name = soil
        ws = wb.create_sheet(title=name)
        # 读取栅格数据
        with rasterio.open(rasterpath) as src:
            # nodata = src.nodata
            # data = src.read(1)
            # profile = src.profile
            # data[data == nodata] = np.nan
            img, _ = mask(src, ya.geometry, crop=True, nodata=np.nan)
            data = img[0]
            data = data[data != src.nodata]
            
        # 统计栅格数据
        dt = pd.DataFrame()
        dt['分级'] = ["一级","二级","三级","四级","五级","全县"]
        # 制图统计
        lst_of_area = []
        lst_of_percentage = []
        total = np.sum(~np.isnan(data))
        #样点统计
        name = feild[name]
        taget = point[name].dropna().astype(float)
        taget[taget == 0] = np.nan
        taget = taget.dropna()
        number = []
        ranges = []
        total_number = taget.count()
        total_range = f'{np.nanmin(taget):.1f}~{np.nanmax(taget):.1f}'
        # 统计数据
        for clas in values[0]:
            if len(clas) == 2:    # 正常情况
                down = clas[0]
                up = clas[1]
                # 栅格统计
                count = np.sum((down<=data) & (data<up))
                # 样点
                filters = taget[(taget >= down) & (taget < up)]
            elif len(clas) == 3:    # 特殊情况
                down = clas[0]
                up = clas[1]
                other = clas[2]
                # 栅格统计
                count = np.sum(((down<=data) & (data<up)) | (data>=other))
                # 样点
                filters = taget[((taget >= down) & (taget < up)) | (taget >= other)]
            elif len(clas) == 4:    # 特殊情况
                down = clas[0]
                up = clas[1]
                other1 = clas[2]
                other2 = clas[3]
                # 栅格统计
                count = np.sum(((down<=data) & (data<up)) | ((data>=other1) & (data<other2)))
                # 样点
                filters = taget[((taget >= down) & (taget < up)) | ((taget >= other1) & (taget < other2))]

            if count == 0:
                lst_of_area.append('-')
                lst_of_percentage.append('-')
            else:
                area = int(count*30*30/666.6667)
                lst_of_area.append(area)
                percentage =(round((count/total),3))*100
                lst_of_percentage.append(percentage) 
                
            if len(filters) == 0:
                ranges.append('-')
                number.append('-')
            else:
                range_value = f'{filters.min():.1f} ~ {filters.max():.1f}'
                number.append(filters.count())
                ranges.append(range_value)
    
        lst_of_area.append(int(total*30*30/666.6667))
        lst_of_percentage.append(100)
        number.append(total_number)
        ranges.append(total_range)
        
        unit = values[1]
        
        dt['面积/亩'] = lst_of_area
        dt['占比/%'] = lst_of_percentage
        dt[f'值域/({unit})'] = ranges
        dt['个数/个'] = number
        # 创建新表格
        ws.append(['土壤三普分级','','','样点统计',''])         #第一行
        # 合并单元格
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=3)
        ws.merge_cells(start_row=1,start_column=4,end_row=1,end_column=5)
        ws.append(list(dt.columns))       #第二行
        # 填充数据
        for excol, dtcol in zip(['A', 'B', 'C', 'D', 'E'], dt.columns):    # 外层循环，循环列
            for i, item in zip(range(3, 9), dt[dtcol]):    # 内层循环，循环列
                ws[f'{excol}{i}'] = item
    
        ws['A9'] = f'全县均值/({unit})'
        ws['A10'] = f'全县中位值/({unit})'
        ws['A11'] = f'全县范围/({unit})'
        ws['B9'] = float(f'{np.nanmean(data):.1f}')    # 求平均值
        ws['B10'] = float(f'{np.nanmedian(data):.1f}')    # 求中位数
        ws['B11'] = f'{np.nanmin(data):.1f}~{np.nanmax(data):.1f}'
    
        # 合并单元格
        ws.merge_cells(start_row=9,start_column=2,end_row=9,end_column=5)
        ws.merge_cells(start_row=10,start_column=2,end_row=10,end_column=5)
        ws.merge_cells(start_row=11,start_column=2,end_row=11,end_column=5)
    
        uniform_width = 20    # 设置列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = uniform_width
            
        # 设置所有单元格居中对齐
        align = Alignment(horizontal='center', vertical='center', wrapText=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
    else:
        print(soil + ' 栅格不存在')
        lack.append(soil)

wb.remove(wb["Sheet"])
wb.save(excel_path)    #保存文件
print('\n表格统计完成')






