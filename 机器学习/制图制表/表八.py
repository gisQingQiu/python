# -*- coding: utf-8 -*-
"""
Created on Mon Sep 22 21:57:50 2025

@author: Qiu
"""

import os
from datetime import datetime
import numpy as np
import pandas as pd
import rasterio
import geopandas as gpd
import openpyxl 
import json
from tqdm import tqdm
from rasterio.mask import mask
from openpyxl.styles import Alignment

start_time = datetime.now()
# 读取数据
with open(r"制表\分级标准.json", encoding='utf-8') as file: 
    classify = json.load(file)
with open(r"制表\字段映射表.json", encoding='utf-8') as file: 
    feild = json.load(file)

soils = r"栅格数据"
polygon = gpd.read_file(r"土地利用现状.shp")

excel_path = r"制表\表九.xlsx"      #表的存放路径
cells = ['A', 'B', 'C', 'D', 'E', 'F']
# 预设的各地类总面积（亩）
area_targets = [303396, 42849, 1440026, 12489]

wb = openpyxl.Workbook()                #创建excel工作簿
ws = wb.active                        #打开一个表格

lack = []

def adjust_areas(pixel_areas, target_area):
    """
    面积平差函数
    将基于像元统计的面积调整到与目标面积相等
    只调整非零值，保持各级别面积的比例关系
    """
    if len(pixel_areas) == 0:
        return []
    
    # 计算当前统计总面积
    current_total = sum(pixel_areas)
    
    # 如果当前总面积为0，无法平差，直接返回原值
    if current_total == 0:
        return pixel_areas
    
    # 计算平差系数
    adjust_ratio = target_area / current_total
    
    # 找出非零值的索引
    non_zero_indices = [i for i, area in enumerate(pixel_areas) if area > 0]
    
    if not non_zero_indices:
        return pixel_areas
    
    # 只对非零值进行平差
    adjusted_areas = pixel_areas.copy()
    
    # 第一次调整：按比例缩放非零值
    for idx in non_zero_indices:
        adjusted_areas[idx] = int(round(pixel_areas[idx] * adjust_ratio))
    
    # 检查调整后的总和
    adjusted_total = sum(adjusted_areas)
    difference = target_area - adjusted_total
    
    # 如果存在微小差异，在最大的非零值上进行修正
    if difference != 0 and non_zero_indices:
        # 找到最大的非零值索引
        max_idx = max(non_zero_indices, key=lambda i: adjusted_areas[i])
        adjusted_areas[max_idx] += difference
    
    return adjusted_areas

for soil, values in tqdm(list(classify.items()), desc='土壤属性'):
    rasterpath = os.path.join(soils, soil+'.tif')
    if os.path.exists(rasterpath):
        # 创建新表格
        name = soil
        ws = wb.create_sheet(title=soil)
        ws.append([f'{soil}含量级别', '分级标准', '单位/亩', '', '', ''])
        ws.append(['', '', '耕地', '园地', '林地', '草地'])
        ws['A8'] = '合计'
        # 合并表格
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)
        ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)
        
        for i, lv in enumerate(["一级","二级","三级","四级","五级"]):
            ws[f'A{i+3}'] = lv
            
        # 存储每个地类的各级别面积（用于平差）
        land_class_areas = {land: [] for land in [1, 2, 3, 4]}
        
        for land in tqdm([1, 2, 3, 4], desc='分类', leave=False):
            target = polygon[polygon['地类']==land]
            
            # 读取栅格数据
            with rasterio.open(rasterpath) as src:
                img, _ = mask(src, target.geometry, crop=True, nodata=np.nan)
                data = img[0]
                data = data[data != src.nodata]
                
                # 计算各地类总面积（用于验证）
                total_pixels = np.sum(~np.isnan(data))
                total_area_pixels = int(round(total_pixels * 30 * 30 / 666.6667))
                ws[f'{cells[land+1]}8'] = total_area_pixels
                
                # 统计各级别面积
                level_areas = []
                for j, clas in enumerate(values[0]):
                    if len(clas) == 2:    # 正常情况
                        down = clas[0]
                        up = clas[1]
                        count = np.sum((down<=data) & (data<up))
                        
                    elif len(clas) == 3:    # 特殊情况
                        down = clas[0]
                        up = clas[1]
                        other = clas[2]
                        count = np.sum(((down<=data) & (data<up)) | (data>=other))
                       
                    elif len(clas) == 4:    # 特殊情况
                        down = clas[0]
                        up = clas[1]
                        other1 = clas[2]
                        other2 = clas[3]
                        count = np.sum(((down<=data) & (data<up)) | ((data>=other1) & (data<other2)))
                    
                    area_mu = int(round(count * 30 * 30 / 666.6667))
                    level_areas.append(area_mu)
                
                # 存储当前地类的各级别面积
                land_class_areas[land] = level_areas
        
        # 平差处理并写入结果
        for land in [1, 2, 3, 4]:
            # 获取当前地类的目标面积
            target_area = area_targets[land-1]  # land从1开始，索引从0开始
            
            # 获取当前地类的各级别面积
            level_areas = land_class_areas[land]
            
            # 进行平差
            adjusted_areas = adjust_areas(level_areas, target_area)
            
            # 写入平差后的结果
            for j, area_value in enumerate(adjusted_areas):
                ws[f'{cells[land+1]}{j+3}'] = area_value
            
            # 更新合计行（应该等于目标面积）
            ws[f'{cells[land+1]}8'] = target_area
        
        # 写入分级标准
        for j, clas in enumerate(values[0]):
            if len(clas) == 2:    # 正常情况
                down = clas[0]
                up = clas[1]
                ws[f'B{j+3}'] = f'{down}-{up}'
            elif len(clas) == 3:    # 特殊情况
                down = clas[0]
                up = clas[1]
                other = clas[2]
                ws[f'B{j+3}'] = f'{down}-{up} & ≥{other}'
            elif len(clas) == 4:    # 特殊情况
                down = clas[0]
                up = clas[1]
                other1 = clas[2]
                other2 = clas[3]
                ws[f'B{j+3}'] = f'{down}-{up} & {other1}-{other2}'
    
        uniform_width = 20    # 设置列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = uniform_width
            
        # 设置所有单元格居中对齐
        align = Alignment(horizontal='center', vertical='center', wrapText=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
    else:
        print(soil + ' 栅格不存在')
        lack.append(soil)

wb.remove(wb["Sheet"])
wb.save(excel_path)    #保存文件
end_time = datetime.now()
print('\n表格统计完成')
print(f'程序用时：{end_time-start_time}')






















