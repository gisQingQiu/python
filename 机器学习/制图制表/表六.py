# -*- coding: utf-8 -*-
"""
Created on Wed Jul  9 09:48:46 2025

@author: Qiu
"""

import numpy as np
import pandas as pd
import geopandas as gpd
import openpyxl
import rasterio
from rasterio.mask import mask
# from glob import glob
from tqdm import tqdm
from openpyxl.styles import Alignment
import os
import json

# 数据路径
# soils = glob(r"栅格数据\*.tif")
soils = r"栅格数据"
ya = gpd.read_file(r"亚类.shp").dropna(subset=['亚类_3']).set_index('亚类_3').to_crs('EPSG:4527')
tu = gpd.read_file(r"土属.shp").dropna(subset=['土属_3']).set_index('土属_3').to_crs('EPSG:4527')
point = gpd.read_file(r"样本点.shp").to_crs('EPSG:4527')
point[(point=='/')|(point=='<空>')] = np.nan
with open(r'制表\亚类土属对应表.json', encoding='utf-8') as file: dic = json.load(file)
with open(r"制表\分级标准.json", encoding='utf-8') as file: claparams = json.load(file)
with open(r"制表\字段映射表.json") as file: feild = json.load(file)

# 创建 Excel 工作簿
excel_path = r"制表\表六.xlsx"      #表的存放路径
wb = openpyxl.Workbook()                #创建excel工作簿
lack = []

for soil in tqdm(claparams.keys(), desc='土壤属性'):
    rasterpath = os.path.join(soils, soil+'.tif')
    if os.path.exists(rasterpath):
        # 创建新表格
        # name = os.path.basename(soil).split('.')[0]
        name = soil
        ws = wb.create_sheet(title=name)
        ws.append(['土壤分级','','样点统计','','', '制图统计', '', ''])         #第一行
        unit = claparams[name][1] if name in claparams.keys() else ''
        name = feild[name]
        # 和并单元格
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2)
        ws.merge_cells(start_row=1,start_column=3,end_row=1,end_column=5)
        ws.merge_cells(start_row=1,start_column=6,end_row=1,end_column=9)
        ws.append(['亚类', '土属', f'均值/({unit})', f'中位值/({unit})', '数量/个', f'范围/({unit})', f'均值/({unit})', '面积/亩'])
        # 样点统计
        # dt = pd.DataFrame(index=['均值/(g/kg)', '中位值/(g/kg)', '数量/个'])
        point_pos = 3
        for key, values in dic.items():
            start_row = point_pos
            for value in values:
                # 按土属提取点数据
                tudf = tu.loc[[value]].copy()
                tudf = gpd.GeoDataFrame(tudf, geometry="geometry")
                tudt = gpd.sjoin(point, tudf, how='inner', predicate='within')
                tudt = tudt[name]
                tudt[tudt==0] = np.nan
                tudt = tudt.dropna().astype(float)
                # if tudt.empty:
                #     dt[value] = ['-']*3
                # else:
                #     dt[value] = [np.nanmean(tudt), np.median(tudt), len(tudt)]
                recodes = ['-']*3 if tudt.empty else [round(np.nanmean(tudt), 1), round(np.median(tudt), 1), tudt.shape[0]]
                # 填入数据
                ws[f'B{point_pos}'] = value
                excel_pos = ['C', 'D', 'E']
                for i, recode in enumerate(recodes):
                    ws[f'{excel_pos[i]}{point_pos}'] = recode
                point_pos += 1
            # 亚类统计
            yadf = ya.loc[[key]].copy()
            yadf = gpd.GeoDataFrame(yadf, geometry="geometry")
            yadt = gpd.sjoin(point, yadf, how='inner', predicate='within')
            yadt = yadt[name]
            yadt[yadt==0] = np.nan
            yadt = yadt.dropna().astype(float)
            # if yadt.empty:
            #     dt[key] = ['-']*3
            # else:
            #    dt[key] = [np.nanmean(yadt), np.median(yadt), len(yadt)]
            recodes = ['-']*3 if yadt.empty else [round(np.nanmean(yadt), 1), round(np.median(yadt), 1), yadt.shape[0]]
            # 填入数据
            ws[f'A{start_row}'] = key
            ws[f'B{point_pos}'] = '合计'
            excel_pos = ['C', 'D', 'E']
            for i, recode in enumerate(recodes):
                ws[f'{excel_pos[i]}{point_pos}'] = recode
            ws.merge_cells(start_row=start_row,start_column=1,end_row=point_pos,end_column=1)
            point_pos += 1
            
        # 制图统计
        # df = pd.DataFrame(index=['范围/(g/kg)', '均值/(g/kg)', '面积/亩'])
        pic_pos = 3
        with rasterio.open(rasterpath) as src:
            for key, values in dic.items():
                for value in values:
                    tudf = tu.loc[[value]].copy()
                    tudf = gpd.GeoDataFrame(tudf, geometry="geometry")
                    tudf = tudf.to_crs(src.crs)
                    img, _ = mask(src, tudf.geometry, crop=True, nodata=np.nan)
                    mask_img = img[0]
                    mask_img = mask_img[mask_img != src.nodata]
                    recodes = [f'{np.nanmin(mask_img):.1f}~{np.nanmax(mask_img):.1f}', f'{np.nanmean(mask_img):.1f}', int((np.sum(~np.isnan(mask_img))*30*30)/666.6667)]
                    # 填入数据
                    excel_pos = ['F', 'G', 'H']
                    for i, recode in enumerate(recodes):
                        ws[f'{excel_pos[i]}{pic_pos}'] = recode
                    pic_pos += 1
                    
                # 亚类统计
                yadf = ya.loc[[key]].copy()
                yadf = gpd.GeoDataFrame(yadf, geometry="geometry")
                yadf = yadf.to_crs(src.crs)
                img, _ = mask(src, yadf.geometry, crop=True, nodata=np.nan)
                mask_img = img[0]
                mask_img = mask_img[mask_img != src.nodata]
                # df[key] = [f'{np.nanmin(mask_img)}~{np.nanmax(mask_img)}', np.nanmean(mask_img), (np.sum(~np.isnan(mask_img))*30*30)/666.6667]
                recodes = [f'{np.nanmin(mask_img):.1f}~{np.nanmax(mask_img):.1f}', round(np.nanmean(mask_img), 1), int((np.sum(~np.isnan(mask_img))*30*30)/666.6667)]
                for i, recode in enumerate(recodes):
                    ws[f'{excel_pos[i]}{pic_pos}'] = recode
                pic_pos += 1
                
        # 全县统计
        full_point = point[name]
        full_point = full_point.replace(0, np.nan)
        full_point = full_point.dropna().astype(float)
        with rasterio.open(rasterpath) as src: 
            img, _ = mask(src, ya.geometry, crop=True, nodata=np.nan)
            mask_img = img[0]
            mask_img = mask_img[mask_img != src.nodata]

        ws.append(['全县', '', round(np.mean(full_point), 1), round(np.median(full_point), 1), full_point.shape[0], f'{np.nanmin(mask_img):.1f}~{np.nanmax(mask_img):.1f}', round(np.nanmean(mask_img), 1), int((np.sum(~np.isnan(mask_img))*30*30)/666.6667)])
        ws.merge_cells(start_row=pic_pos,start_column=1,end_row=pic_pos,end_column=2)
        
        uniform_width = 20    # 设置列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = uniform_width
        
        # 设置所有单元格居中对齐
        align = Alignment(horizontal='center', vertical='center', wrapText=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align
    else:
        print(soil + ' 栅格不存在')
        lack.append(soil)
        
wb.remove(wb["Sheet"])
wb.save(excel_path)














