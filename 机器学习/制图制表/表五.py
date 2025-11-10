# -*- coding: utf-8 -*-
"""
Created on Fri Mar 14 20:24:49 2025

@author: Qiu
"""

import os
import json
import numpy as np
import pandas as pd
import geopandas as gpd
import rasterio
from rasterio.mask import mask
from tqdm import tqdm
import openpyxl
from glob import glob
from openpyxl.styles import Alignment

def get_range(range_str):
    # 分割范围字符串
    min_val, max_val = range_str.split('~')
    return float(min_val), float(max_val)

with open(r"制表\分级标准.json", encoding='utf-8') as file: classify = json.load(file)
with open(r"制表\字段映射表.json") as file: feild = json.load(file)
soils = r"栅格数据"
point = gpd.read_file(r"D样本点.shp").to_crs('EPSG:4527')
point[(point=='/')|(point=='<空>')] = np.nan
polygon = gpd.read_file(r"土地利用现状.shp")
dic = {'1':'水田', '2':'水浇地', '3':'旱地', '4':'园地', '5':'林地', '6':'草地', '7':'水体', '8':'建设用地', '9':'其他'}

point[point == 0] = np.nan
point[point == '0'] = np.nan
point['分类'] = [dic[i] for i in list(point['LUCC_1'])]
# 创建 Excel 工作簿
excel_path = r"制表\表五.xlsx" 
wb = openpyxl.Workbook()
ws = wb.active 

for soil, values in tqdm(classify.items(), desc='土壤属性'):
    rasterpath = os.path.join(soils, soil+'.tif')
    unit = values[1] if soil in classify.keys() else ''
    if os.path.exists(rasterpath):
        # 创建新表格
        ws = wb.create_sheet(title=soil)
        name = feild[soil]
    else:
        continue
    
    # 样点统计
    dt = pd.DataFrame(index=[f'均值/{unit}', f'范围/{unit}', '数量/个'])
    for soil_type in dic.values():
        pt_type = point.loc[point['分类'] == soil_type]
        feature = pt_type[name].dropna().astype(float)
        if feature.shape[0] == 0:
            dt[soil_type] = ['-']*3
        else:
            dt[soil_type] = [round(np.nanmean(feature), 1), f'{np.nanmin(feature):.1f}~{np.nanmax(feature):.1f}', feature.shape[0]]
    
    ranges = dt.loc[f'范围/{unit}', ['水田', '旱地']]

    ranges = [get_range(i) for i in ranges]
    ranges = [val for pair in ranges for val in pair]
    dt['合计'] = [round(np.nanmean(dt.loc[f'均值/{unit}', ['水田', '旱地']]), 1), f'{np.nanmin(ranges):.1f}~{np.nanmax(ranges):.1f}', round(np.sum(dt.loc['数量/个', ['水田', '旱地']]), 1)]
    dt = dt[['水田', '水浇地', '旱地', '合计', '园地', '林地', '草地', '水体', '建设用地', '其他']]
    dt = dt.T
 
    # 制图统计
    df = pd.DataFrame(index=[f'均值/{unit}', '面积/亩'])
    polygon['分类'] = [dic[i] for i in list(polygon['LUCC'])]
    with rasterio.open(rasterpath) as src:
        # 统一投影
        polygon_crs = polygon.crs
        src_crs = src.crs
        if polygon_crs != src_crs:
            polygon = polygon.to_crs(src.crs)
        nodata = src.nodata
        data = src.read(1).astype(float)
        data[data==nodata] = np.nan
        # 统计
        for soil_type, geometry in tqdm(zip(polygon['分类'], polygon['geometry']),total=polygon.shape[0], desc='栅格统计', leave=False):
            mask_image, _ = mask(src, [geometry], crop=True)
            mask_data = mask_image[0]
            mask_data = mask_data[mask_data != nodata]
            df[soil_type] = [np.nanmean(mask_data), int((np.sum(~np.isnan(mask_data))*30*30)/666.6667)]
        df['合计'] = [np.nanmean(df.loc[f'均值/{unit}', ['水田', '水浇地', '旱地']]), np.nansum(df.loc['面积/亩', ['水田', '水浇地', '旱地']])]
        df = df[['水田', '水浇地', '旱地', '合计', '园地', '林地', '草地', '水体', '建设用地', '其他']]
        df = df.T

    # 写入表头
    ws.append(['土地利用类型', '', '样点统计', '', '', '制图统计', ''])  # 第一行
    ws.append(['一级', '二级', f'均值/{unit}', f'范围/{unit}', '数量/个', f'均值/{unit}', '面积/亩'])
    # 合并表头单元格
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)  # "土地利用类型"
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)  # "样点统计"
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)  # "制图统计"
    
    ws['A3'] = '耕地'
    ws.merge_cells(start_row=3, start_column=1, end_row=6, end_column=1)
    
    for row, value in enumerate(['水田', '水浇地', '旱地', '合计'], start=3):  # 从第 3 行开始
        ws[f'B{row}'] = value
        
    for row, value in enumerate(['园地', '林地', '草地', '水体', '建设用地', '其他'], start=7):
        ws[f'A{row}'] = value
    for i in range(7, 13):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    
    # 填充样点统计数据
    for row, land_type in enumerate(['水田', '水浇地', '旱地', '合计', '园地', '林地', '草地', '水体', '建设用地', '其他'], start=3):
        if land_type in dt.index:
            try:
                ws[f'C{row}'] = dt.loc[land_type, f'均值/{unit}']
            except:
                ws[f'C{row}'] = dt.loc[land_type, f'均值/{unit}']
            ws[f'D{row}'] = dt.loc[land_type, f'范围/{unit}']
            ws[f'E{row}'] = dt.loc[land_type, '数量/个']
    
    # 填充制图统计数据
    for row, land_type in enumerate(['水田', '水浇地', '旱地', '合计', '园地', '林地', '草地', '水体', '建设用地', '其他'], start=3):
        if land_type in df.index:
            ws[f'F{row}'] = round(df.loc[land_type, f'均值/{unit}'], 1)
            ws[f'G{row}'] = df.loc[land_type, '面积/亩']
    allpoint = point[name].dropna().astype(float)
    ws.append(['全县', '', round(np.nanmean(allpoint.values), 1), f'{np.nanmin(allpoint):.1f}~{np.nanmax(allpoint):.1f}', allpoint.shape[0], round(np.nanmean(data), 1), int((np.sum(~np.isnan(data))*30*30)/666.6667)])
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=2)
    
    uniform_width = 15    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = uniform_width
        
    # 设置所有单元格居中对齐
    align = Alignment(horizontal='center', vertical='center', wrapText=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align
     
wb.remove(wb["Sheet"])   
wb.save(excel_path)
print('-'*60)
print('完成所有土壤属性统计')


